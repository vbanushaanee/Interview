import pandas as pd
from openpyxl import Workbook,load_workbook
import re
 
#first i have converted data.txt to data.xlsx to get results in a easy way

# As data not in proper format so made it as proper 

file="data.txt"
def reading_data(file):
    with open(file,"r") as f:
        data=f.readlines()
        res=[]
        for i in data:
            res.append(i)
    
    actual_data=[]
        
    per1 = pd.date_range(start ='08-01-2022', end ='12-31-2022')

    for i in res:
        kk=i.split(",")
        if kk[1] in per1:
            actual_data.append(kk)
            
    header_li=["ID","Date","Description/Narration","Value date","ChqJRef.No.","Debit(Dr.)","Creditf(Cr.)","Balance"]
    actual_data.insert(0,header_li)
    
    wb = Workbook()
    ws = wb.active

    for i in range(1,len(res)):
        dd=res[i]
        kk=dd.split(",")
        for j in range(1,len(kk)):
            
            ws.cell(row=i,column=j,value=kk[j])
      
      

    wb.save("data.xlsx")
    
    
    return actual_data
   
actual_data=reading_data(file)   
    


wb1=load_workbook("data.xlsx")
wb1.create_sheet("bank_trans")
ws1 = wb1["bank_trans"]
for i in range(0,len(actual_data)):
    dd=actual_data[i]
    
    for j in range(0,len(dd)):
        
        ws1.cell(row=i+1,column=j+1,value=dd[j])

wb1.save("data.xlsx")


df = pd.read_excel('data.xlsx', sheet_name = "bank_trans")

#top 10 records of credit
sdf=df.sort_values("Creditf(Cr.)",ascending = False).head(10)

#print(sdf)

#output
'''
ID          Date                              Description/Narration    Value date  ... Creditf(Cr.)       Balance Unnamed: 8 Unnamed: 9
8    17   23 Aug 2022  UPI/DR/223576017763/ISHWAR LAL DABl/PYTM/91982...   23 Aug 2022  ...       793.95           NaN         \n        NaN
157  14  09 Dec 2022   UPI/CR/234317998040/MAHENDRASINGH SOLANKI/BKID...  09 Dec 2022   ...         600    10607.33 \n        NaN        NaN
124  12  15 Nov 2022   UPI/CR/231997183369/SUNIL BAIRAGI/SBIN/0000003...  15 Nov 2022   ...       50000   190009.33 \n        NaN        NaN
116   4  04 Nov 2022   UPI/CR/230844256028/SUNIL BAIRAGI/SBIN/0000003...  04 Nov 2022   ...        5000     5003.33 \n        NaN        NaN
69   19  15 Sep 2022   UPI/CR/225833624874/SUNIL BAIRAGI/SBIN/0000003...  15 Sep 2022   ...         500      502.95 \n        NaN        NaN
100  19  17 Oct 2022   UPI/CR/229053391278/SUNIL BAIRAGI/SBIN/0000003...  17 Oct 2022   ...         500         500 \n        NaN        NaN
20    1  01 Sep 2022                            MONTHLY INTEREST PAYOUT   31 Aug 2022   ...           5    13413.95 \n        NaN        NaN
71   21  15 Sep 2022   PI/DR-REV/225883269207/Q378887034/YBL/Q3788870...  15 Sep 2022   ...          42      502.95 \n        NaN        NaN
123  11  15 Nov 2022                              CASH DEP FREE GANJ UJ   15 Nov 2022   ...       40000   140009.33 \n        NaN        NaN
138  26  02 Dec 2022                                       SALARY NOV22   02 Dec 2022   ...       35676    42894.33 \n        NaN        NaN

[10 rows x 10 columns]

'''
  

  
#top 10 records of debit
sdf_debit=df.sort_values("Debit(Dr.)",ascending = False).head(10)
#print(sdf_debit)
#output
'''
 ID          Date                              Description/Narration    Value date  ... Creditf(Cr.)       Balance Unnamed: 8 Unnamed: 9
2    11   22 Aug 2022  UPI/DR/223492496549AJEEVAN SINGH4NDB/2O100451 ...   22 Aug 2022  ...           30       1081.95        NaN         \n
72   22  16 Sep 2022   UPI/DR/225967576958/POONAM GROCERY/@FRE/FCBIZW...  16 Sep 2022   ...           -      410.95 \n        NaN        NaN
81   31  04 Oct 2022   UPI/DR/227786967153/SUNIL BAIRAGI/SBIN/0000003...  04 Oct 2022   ...           -     4648.95 \n        NaN        NaN
36   17  01 Sep 2022   UPI/DR/224473034542/SHREE MAHAKAL IRANA/YESB/1...  01 Sep 2022   ...           -     7842.95 \n        NaN        NaN
22    3  01 Sep 2022   UPI/DR/224435085252/HEMRAJ RAIKWAR/YESB/002261...  01 Sep 2022   ...           -    13323.95 \n        NaN        NaN
93   12  07 Oct 2022   UPI/DR/228002613781/MULCHANDSOKUNDAMALNA/YESB/...  07 Oct 2022   ...           -       68.95 \n        NaN        NaN
60   10  07 Sep 2022   UPI/DR/225073124022/JITENDRA JAIN so RAJENDRA ...  07 Sep 2022   ...           -      455.95 \n        NaN        NaN
121   9  11 Nov 2022   UPI/DR/231511712089/CHETAN/YESB/00226110000002...  11 Nov 2022   ...           -        9.33 \n        NaN        NaN
16   25   31 Aug 2022  UPI/DR/224371592649A/INOD PRAJAPATI SORAMPRASA...   31 Aug 2022  ...          NaN      14819.95        NaN         \n
127  15  18 Nov 2022   UPI/DR/232262572656/VIJAY KHUSHLANI/@YBL/Q4829...  18 Nov 2022   ...           -   189859.33 \n        NaN        NaN

[10 rows x 10 columns]
'''


#---------similar transactions----------

new = df.groupby([df.duplicated(["Description/Narration"])])

print("new:",list(new))
#output
'''

[(False,      ID          Date                              Description/Narration    Value date  ... Creditf(Cr.)      Balance Unnamed: 8 Unnamed: 9
0     9   22 Aug 2022   IMPS-223407656547-YES8-XXXXXXXXXXX0379-PENNYDROP   22 Aug 2022  ...            1      1141.95        NaN         \n
1    10   22 Aug 2022  UPI/DR/223490715044/PUNAM GROSHARI HOME/UTlB/1...   22 Aug 2022  ...          NaN      1111.95        NaN         \n
2    11   22 Aug 2022  UPI/DR/223492496549AJEEVAN SINGH4NDB/2O100451 ...   22 Aug 2022  ...           30      1081.95        NaN         \n
3    12   22 Aug 2022  UPI/DR/22342022B785/COFFEE CLUB/PYTM/197442010...   22 Aug 2022  ...          NaN       981.95        NaN         \n
4    13   22 Aug 2022  UPI/DR/223407646216/GOVINDA SONS/PYTM/19744201...   22 Aug 2022  ...          NaN       957.95        NaN         \n
..   ..           ...                                                ...           ...  ...          ...          ...        ...        ...
155  12  05 Dec 2022   UPI/DR/233966239432/RAVINDRA: SINGH JAISWAL/PY...  05 Dec 2022   ...           -       7.33 \n        NaN        NaN
156  13  09 Dec 2022   UPICR/234397055237/MAHENDRA SINGH SOLANKI/BKID...  09 Dec 2022   ...       10000   10007.33 \n        NaN        NaN
157  14  09 Dec 2022   UPI/CR/234317998040/MAHENDRASINGH SOLANKI/BKID...  09 Dec 2022   ...         600   10607.33 \n        NaN        NaN
158  15  09 Dec 2022   UPI/DR/234390604766/DULICHNDR JAUBL/2211717843...  09 Dec 2022   ...           -    5607.33 \n        NaN        NaN
159  16  09 Dec 2022   PI/DR/234393862247/RAKESH/BARB/57710100002280/...  09 Dec 2022   ...           -       7.33 \n        NaN        NaN

[155 rows x 10 columns]), (True,      ID          Date                              Description/Narration    Value date  ... Creditf(Cr.)     Balance Unnamed: 8 Unnamed: 9
85    4  05 Oct 2022   9001250527925404 (DRAWDOWN FROM CASA)-SUNIL BA...  05 Oct 2022   ...           -    711.95 \n        NaN        NaN
103  22  18 Oct 2022                      DEBIT CARD FEE FY22'23 XX3434   18 Oct 2022   ...           -    187.95 \n        NaN        NaN
119   7  05 Nov 2022   9001250527925404 (DRAWDOWN FROM CASA)-SUNIL BA...  05 Nov 2022   ...           -   2437.33 \n        NaN        NaN
137  25  01 Dec 2022                            MONTHLY INTEREST PAYOUT   30 Nov 2022   ...         217   7218.33 \n        NaN        NaN
152   9  05 Dec 2022   9001250527925404 (DRAWDOWN FROM CASA)-SUNIL BA...  05 Dec 2022   ...           -   2230.33 \n        NaN        NaN

[5 rows x 10 columns])]
'''


#-----------similar transactions mean values ---------
print(new.describe())

#output
'''
          ID
       count       mean       std  min  25%   50%   75%   max
False  155.0  15.632258  8.655769  1.0  9.0  15.0  23.0  31.0
True     5.0  13.400000  9.449868  4.0  7.0   9.0  22.0  25.0
'''














